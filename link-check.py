import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime

# ========== Configuration ==========
SEARCH_ROOT = r'C:\Projects'  # Root directory where all local repos are stored
input_file = r"C:\Users\v-bowenyang\Desktop\Invalid-link-check.xlsx"

# Create output directory on Desktop
desktop = os.path.join(os.path.expanduser("~"), "Desktop")
output_dir = os.path.join(desktop, "Link-checker")
os.makedirs(output_dir, exist_ok=True)

# Generate timestamped output file name
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
output_file = os.path.join(output_dir, f"Check_Result_{timestamp}.xlsx")

# ========== Load input Excel ==========
df = pd.read_excel(input_file)
df.columns = df.columns.str.strip()  # Clean up column names

results = []

# ========== Process each row ==========
for index, row in df.iterrows():
    link_path = str(row['Broken link ID']).strip()
    github_url = str(row['Source URL']).strip()

    if '/blob/' in github_url:
        try:
            # Extract repo name and relative path from GitHub URL
            repo_info = github_url.split('github.com/')[1].split('/blob/', 1)
            repo_parts = repo_info[0].split('/')
            repo_name = repo_parts[1]
            relative_path = repo_info[1].split('/', 1)[1]

            # Construct full local path using SEARCH_ROOT and repo name
            local_repo_path = os.path.join(SEARCH_ROOT, repo_name)
            local_path = os.path.join(local_repo_path, relative_path.replace('/', os.sep))

        except Exception:
            results.append('Path parsing failed')
            continue
    else:
        results.append('Invalid URL')
        continue

    print(f'[{index+1}/{len(df)}] Checking file: {local_path}')

    try:
        if not os.path.exists(local_path):
            results.append('File not found locally')
            continue

        with open(local_path, 'r', encoding='utf-8') as f:
            content = f.read()
            if link_path in content:
                results.append('')  # Leave empty if link is found
            else:
                results.append('NF')  # Not Found

    except Exception as e:
        print(f'❌ Error: {e}')
        results.append('Read failed')

# ========== Save results ==========
df['Check Result'] = results
df.to_excel(output_file, index=False)

# ========== Beautify Excel Output ==========
wb = load_workbook(output_file)
ws = wb.active

# Set font to DengXian and adjust column widths
font = Font(name='DengXian', size=11)
for row in ws.iter_rows():
    for cell in row:
        cell.font = font

for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            val = str(cell.value)
            if len(val) > max_length:
                max_length = len(val)
        except:
            pass
    adjusted_width = max_length + 2
    ws.column_dimensions[col_letter].width = adjusted_width

wb.save(output_file)

# ========== Open Excel file ==========
print(f'\n✅ Check completed. Results saved and formatted:\n{output_file}')

try:
    os.startfile(output_file)
except Exception as e:
    print(f"⚠️ Failed to open the output file automatically: {e}")
